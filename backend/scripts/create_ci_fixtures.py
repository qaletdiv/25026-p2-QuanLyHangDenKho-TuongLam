"""
create_ci_fixtures.py
=====================
Generates 7 CI (Commercial Invoice) fixture Excel files for the
tentree Supply Chain Portal test suite, plus a CI_FIXTURES.md
documentation file.

Output directory: ../data/ (relative to this script's location)
Parser format expected by: backend/services/ciParser.js

Run:
    python backend/scripts/create_ci_fixtures.py

Requirements:
    openpyxl  (auto-installed if missing)
"""

import sys
import subprocess
import os
from pathlib import Path
from datetime import date

# ── 0. Ensure openpyxl is available ─────────────────────────────────────────

try:
    import openpyxl
except ImportError:
    print("openpyxl not found — installing...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 1. Paths ─────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent          # backend/scripts/
DATA_DIR   = SCRIPT_DIR.parent / "data"               # backend/data/
DATA_DIR.mkdir(parents=True, exist_ok=True)

# ── 2. Shared style helpers ──────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
LABEL_FONT   = Font(bold=True, size=10)
DATA_FONT    = Font(size=10)
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")   # light blue

thin = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header_row(ws, row, cols):
    """Apply dark-blue header styling to a row for the given column range."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def style_section_row(ws, row, cols):
    """Apply light-blue fill to a PO summary data row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = SECTION_FILL
        cell.font   = LABEL_FONT
        cell.border = THIN_BORDER


def style_data_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    """widths: list of (col_letter, width) tuples."""
    for letter, width in widths:
        ws.column_dimensions[letter].width = width


# ── 3. Core workbook builder ─────────────────────────────────────────────────

def build_ci_workbook(invoice_number, invoice_date, po_summary_rows, line_items):
    """
    Constructs a workbook matching the ciParser.js fixed-format layout:

        Row 1    : Title banner
        Row 2    : invoice_number → B2 | invoice_date → D2
        Row 3    : Section label "PO Shipping Summary" (merged A3:E3)
        Rows 4–8 : PO summary block — col A=po_number, B=shipped_qty,
                   C=cartons, D=weight_kg, E=cbm (one row per PO, up to 5)
        Row 9    : blank separator
        Row 10   : SKU line-item header labels (ignored by parser)
        Row 11+  : SKU line items — col A=sku_code, B=description, C=qty,
                   D=unit_price, E=total_price, F=weight_kg, G=cbm
                   (parser stops on first empty col A)

    po_summary_rows : list of (po_number, shipped_qty, cartons, weight_kg, cbm)
    line_items      : list of (sku_code, description, qty, unit_price, weight_kg, cbm)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Commercial Invoice"

    # ── Row 1: Title banner ───────────────────────────────────────────────────
    ws["A1"] = "COMMERCIAL INVOICE"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # ── Row 2: Invoice metadata ───────────────────────────────────────────────
    ws["A2"] = "Invoice No:"
    ws["A2"].font = LABEL_FONT
    ws["B2"] = invoice_number                          # parser reads B2
    ws["B2"].font = Font(bold=True, size=10, color="C00000")
    ws["C2"] = "Invoice Date:"
    ws["C2"].font = LABEL_FONT
    ws["D2"] = invoice_date                            # parser reads D2
    ws["D2"].number_format = "YYYY-MM-DD"
    ws["D2"].font = Font(bold=True, size=10)

    # ── Row 3: Section label for PO block ────────────────────────────────────
    ws["A3"] = "PO Shipping Summary"
    ws["A3"].font = Font(bold=True, size=10, color="1F4E79")
    ws.merge_cells("A3:E3")

    # ── Rows 4–8: PO summary block ────────────────────────────────────────────
    # Parser reads rows 4–8; skips any row where col A is empty.
    for offset, (po_num, s_qty, cartons, wt, cbm) in enumerate(po_summary_rows):
        r = 4 + offset   # rows 4, 5, 6, 7, 8
        ws.cell(row=r, column=1, value=po_num)
        ws.cell(row=r, column=2, value=s_qty)
        ws.cell(row=r, column=3, value=cartons)
        ws.cell(row=r, column=4, value=wt)
        ws.cell(row=r, column=5, value=cbm)
        style_section_row(ws, r, 5)

    # ── Row 9: Blank separator (col A is empty — parser skips it) ────────────
    # No write needed; row is blank by default.

    # ── Row 10: SKU line-item header labels ───────────────────────────────────
    headers = [
        "SKU Code", "Description", "Qty",
        "Unit Price (USD)", "Total Price (USD)", "Weight (kg)", "CBM",
    ]
    for col, label in enumerate(headers, start=1):
        ws.cell(row=10, column=col, value=label)
    style_header_row(ws, 10, len(headers))

    # ── Rows 11+: SKU line items ──────────────────────────────────────────────
    for row_offset, (sku, desc, qty, unit_price, wt, cbm) in enumerate(line_items):
        r = 11 + row_offset
        total_price = round(qty * unit_price, 2)
        ws.cell(row=r, column=1, value=sku)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=qty)
        ws.cell(row=r, column=4, value=unit_price)
        ws.cell(row=r, column=5, value=total_price)
        ws.cell(row=r, column=6, value=wt)
        ws.cell(row=r, column=7, value=cbm)
        style_data_row(ws, r, 7)
        ws.cell(row=r, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5).number_format = '"$"#,##0.00'

    # ── Column widths ─────────────────────────────────────────────────────────
    set_col_widths(ws, [
        ("A", 28), ("B", 36), ("C", 10),
        ("D", 17), ("E", 19), ("F", 12), ("G", 10),
    ])

    return wb


# ── 4. Shared SKU / PO source data ───────────────────────────────────────────

# Each tuple: (sku_code, description, expected_qty, unit_price, wt_per_unit_kg, cbm_per_unit)
PO001_SKUS = [
    ("TEN7101-FGN-XS", "FW26 Fleece Jacket — Forest Green XS",  800,  16.50, 0.20, 0.0025),
    ("TEN7101-FGN-S",  "FW26 Fleece Jacket — Forest Green S",  1400,  16.50, 0.20, 0.0025),
    ("TEN7101-FGN-M",  "FW26 Fleece Jacket — Forest Green M",  2000,  16.50, 0.20, 0.0025),
    ("TEN7101-FGN-L",  "FW26 Fleece Jacket — Forest Green L",  2000,  16.50, 0.20, 0.0025),
    ("TEN7101-FGN-XL", "FW26 Fleece Jacket — Forest Green XL", 1600,  16.50, 0.20, 0.0025),
]  # sum of expected_qty = 7800

PO002_SKUS = [
    ("TEN7102-NVY-XS", "FW26 Hoodie — Navy XS",  500, 18.00, 0.22, 0.0028),
    ("TEN7102-NVY-S",  "FW26 Hoodie — Navy S",  1000, 18.00, 0.22, 0.0028),
    ("TEN7102-NVY-M",  "FW26 Hoodie — Navy M",  1500, 18.00, 0.22, 0.0028),
]  # sum of expected_qty = 3000

INVOICE_DATE = date(2026, 4, 15)


def li(sku_code, description, qty, unit_price, wt_per_unit, cbm_per_unit):
    """Build a line-item tuple, computing total weight and CBM from per-unit values."""
    return (
        sku_code,
        description,
        qty,
        unit_price,
        round(qty * wt_per_unit, 2),    # weight_kg total
        round(qty * cbm_per_unit, 4),   # cbm total
    )


def li_from_source(src_tuple, qty_override=None):
    """Build a line-item tuple from a PO source row, optionally overriding qty."""
    sku, desc, expected_qty, unit_price, wt_per_unit, cbm_per_unit = src_tuple
    qty = qty_override if qty_override is not None else expected_qty
    return li(sku, desc, qty, unit_price, wt_per_unit, cbm_per_unit)


# ── 5. Define all 7 scenarios ─────────────────────────────────────────────────

# Each entry: (filename, workbook, metadata_dict_for_md)
scenarios = []


# ─────────────────────────────────────────────────────────────────────────────
# S1: Full match — all 5 PO-001 SKUs at exact expected qty
# ─────────────────────────────────────────────────────────────────────────────
s1_items = [li_from_source(s) for s in PO001_SKUS]

s1_wb = build_ci_workbook(
    invoice_number  = "CI-S1-FULL-001",
    invoice_date    = INVOICE_DATE,
    po_summary_rows = [("PO-FW26-001", 7800, 390, 1560.0, 19.50)],
    line_items      = s1_items,
)
scenarios.append((
    "ci_s1_full_match.xlsx",
    s1_wb,
    {
        "title": "S1 — Full Match (single PO)",
        "desc": (
            "All 5 PO-FW26-001 SKUs shipped at exact expected quantities. "
            "Zero over/under. Every CI SKU matches a PO line item."
        ),
        "pos": ["PO-FW26-001"],
        "skus": [
            ("TEN7101-FGN-XS",   800, "matched"),
            ("TEN7101-FGN-S",   1400, "matched"),
            ("TEN7101-FGN-M",   2000, "matched"),
            ("TEN7101-FGN-L",   2000, "matched"),
            ("TEN7101-FGN-XL",  1600, "matched"),
        ],
        "parse": {
            "total_shipped_qty": 7800,
            "matched_skus":      5,
            "unmatched_skus":    0,
            "over_shipped_skus": 0,
        },
    },
))


# ─────────────────────────────────────────────────────────────────────────────
# S2: Partial match — 3 SKUs (XS, S, M), all matched, qty < PO expected
# ─────────────────────────────────────────────────────────────────────────────
s2_items = [li_from_source(s) for s in PO001_SKUS[:3]]   # XS + S + M = 4200

s2_wb = build_ci_workbook(
    invoice_number  = "CI-S2-PARTIAL-001",
    invoice_date    = INVOICE_DATE,
    po_summary_rows = [("PO-FW26-001", 4200, 210, 840.0, 10.50)],
    line_items      = s2_items,
)
scenarios.append((
    "ci_s2_partial_match.xlsx",
    s2_wb,
    {
        "title": "S2 — Partial Match (single PO, qty < expected)",
        "desc": (
            "Only 3 of 5 PO-FW26-001 SKUs included (XS, S, M). "
            "Total shipped 4,200 vs 7,800 expected. All CI SKUs match the PO. "
            "L and XL have 0 shipped — their remaining_qty equals expected_qty."
        ),
        "pos": ["PO-FW26-001"],
        "skus": [
            ("TEN7101-FGN-XS",  800, "matched"),
            ("TEN7101-FGN-S",  1400, "matched"),
            ("TEN7101-FGN-M",  2000, "matched"),
        ],
        "parse": {
            "total_shipped_qty": 4200,
            "matched_skus":      3,
            "unmatched_skus":    0,
            "over_shipped_skus": 0,
        },
    },
))


# ─────────────────────────────────────────────────────────────────────────────
# S3: Partial + unmatched — some SKUs not on PO
# ─────────────────────────────────────────────────────────────────────────────
s3_items = [
    li("TEN7101-FGN-XS",      "FW26 Fleece Jacket — Forest Green XS",       800, 16.50, 0.20, 0.0025),
    li("TEN7101-FGN-S",       "FW26 Fleece Jacket — Forest Green S",        1400, 16.50, 0.20, 0.0025),
    li("TEN7101-FGN-UNKNOWN", "FW26 Fleece Jacket — Forest Green UNKNOWN",  1600, 16.50, 0.20, 0.0025),
]

s3_wb = build_ci_workbook(
    invoice_number  = "CI-S3-PARTIAL-UNM-001",
    invoice_date    = INVOICE_DATE,
    po_summary_rows = [("PO-FW26-001", 3800, 190, 760.0, 9.50)],
    line_items      = s3_items,
)
scenarios.append((
    "ci_s3_partial_some_unmatched.xlsx",
    s3_wb,
    {
        "title": "S3 — Partial + Unmatched SKUs (single PO)",
        "desc": (
            "2 matched SKUs (XS=800, S=1400) plus 1 SKU that does not exist on "
            "PO-FW26-001 (TEN7101-FGN-UNKNOWN=1600). "
            "Total CI qty = 3,800 vs PO expected = 7,800."
        ),
        "pos": ["PO-FW26-001"],
        "skus": [
            ("TEN7101-FGN-XS",       800, "matched"),
            ("TEN7101-FGN-S",       1400, "matched"),
            ("TEN7101-FGN-UNKNOWN", 1600, "UNMATCHED — not on PO"),
        ],
        "parse": {
            "total_shipped_qty": 3800,
            "matched_skus":      2,
            "unmatched_skus":    1,
            "over_shipped_skus": 0,
        },
    },
))


# ─────────────────────────────────────────────────────────────────────────────
# S4: Overbooking + unmatched — qty > PO expected on some SKUs
# ─────────────────────────────────────────────────────────────────────────────
s4_items = [
    li("TEN7101-FGN-XS",    "FW26 Fleece Jacket — Forest Green XS",     800, 16.50, 0.20, 0.0025),  # exact
    li("TEN7101-FGN-S",     "FW26 Fleece Jacket — Forest Green S",      2000, 16.50, 0.20, 0.0025),  # over by 600
    li("TEN7101-FGN-M",     "FW26 Fleece Jacket — Forest Green M",      4000, 16.50, 0.20, 0.0025),  # over by 2000
    li("TEN7101-FGN-GHOST", "FW26 Fleece Jacket — Forest Green GHOST",  2200, 16.50, 0.20, 0.0025),  # not on PO
]

s4_wb = build_ci_workbook(
    invoice_number  = "CI-S4-OVERBOOK-001",
    invoice_date    = INVOICE_DATE,
    po_summary_rows = [("PO-FW26-001", 9000, 450, 1800.0, 22.50)],
    line_items      = s4_items,
)
scenarios.append((
    "ci_s4_overbooking_mixed.xlsx",
    s4_wb,
    {
        "title": "S4 — Overbooking + Unmatched (single PO)",
        "desc": (
            "Total CI qty = 9,000 vs PO expected = 7,800. "
            "XS matched exactly (800). S over-shipped by 600 (2000 vs 1400 expected). "
            "M over-shipped by 2000 (4000 vs 2000 expected). "
            "TEN7101-FGN-GHOST=2200 does not exist on PO-FW26-001."
        ),
        "pos": ["PO-FW26-001"],
        "skus": [
            ("TEN7101-FGN-XS",     800, "matched"),
            ("TEN7101-FGN-S",     2000, "OVER-SHIPPED — expected 1,400"),
            ("TEN7101-FGN-M",     4000, "OVER-SHIPPED — expected 2,000"),
            ("TEN7101-FGN-GHOST", 2200, "UNMATCHED — not on PO"),
        ],
        "parse": {
            "total_shipped_qty": 9000,
            "matched_skus":      1,
            "unmatched_skus":    1,
            "over_shipped_skus": 2,
        },
    },
))


# ─────────────────────────────────────────────────────────────────────────────
# S5: All PO SKUs covered + 2 extra CI SKUs not on PO
# ─────────────────────────────────────────────────────────────────────────────
s5_items = (
    [li_from_source(s) for s in PO001_SKUS]   # 7800 — all at exact expected qty
    + [
        li("TEN7101-EXTRA-001", "FW26 Fleece Jacket — Extra Style 001", 1000, 16.50, 0.20, 0.0025),
        li("TEN7101-EXTRA-002", "FW26 Fleece Jacket — Extra Style 002", 1000, 16.50, 0.20, 0.0025),
    ]
)

s5_wb = build_ci_workbook(
    invoice_number  = "CI-S5-EXTRA-SKU-001",
    invoice_date    = INVOICE_DATE,
    po_summary_rows = [("PO-FW26-001", 9800, 490, 1960.0, 24.50)],
    line_items      = s5_items,
)
scenarios.append((
    "ci_s5_overbooking_extra_skus.xlsx",
    s5_wb,
    {
        "title": "S5 — All PO SKUs + Extra CI SKUs (single PO, over-shipped)",
        "desc": (
            "All 5 PO-FW26-001 SKUs at exact expected quantities (7,800 total), "
            "PLUS 2 extra SKUs not present on the PO "
            "(TEN7101-EXTRA-001=1000, TEN7101-EXTRA-002=1000). "
            "Total CI = 9,800 vs PO expected = 7,800."
        ),
        "pos": ["PO-FW26-001"],
        "skus": [
            ("TEN7101-FGN-XS",     800, "matched"),
            ("TEN7101-FGN-S",     1400, "matched"),
            ("TEN7101-FGN-M",     2000, "matched"),
            ("TEN7101-FGN-L",     2000, "matched"),
            ("TEN7101-FGN-XL",    1600, "matched"),
            ("TEN7101-EXTRA-001", 1000, "UNMATCHED — not on PO"),
            ("TEN7101-EXTRA-002", 1000, "UNMATCHED — not on PO"),
        ],
        "parse": {
            "total_shipped_qty": 9800,
            "matched_skus":      5,
            "unmatched_skus":    2,
            "over_shipped_skus": 0,
        },
    },
))


# ─────────────────────────────────────────────────────────────────────────────
# S1 Multi-PO: Both PO-001 and PO-002, all SKUs at exact expected qty
# ─────────────────────────────────────────────────────────────────────────────
s1mp_items = (
    [li_from_source(s) for s in PO001_SKUS]   # 7800
    + [li_from_source(s) for s in PO002_SKUS]  # 3000
)

s1mp_wb = build_ci_workbook(
    invoice_number  = "CI-S1MP-MULTI-001",
    invoice_date    = INVOICE_DATE,
    po_summary_rows = [
        ("PO-FW26-001", 7800, 390, 1560.0, 19.50),
        ("PO-FW26-002", 3000, 150,  660.0,  8.40),
    ],
    line_items = s1mp_items,
)
scenarios.append((
    "ci_s1_multi_po.xlsx",
    s1mp_wb,
    {
        "title": "S1 Multi-PO — Full Match (PO-FW26-001 + PO-FW26-002)",
        "desc": (
            "All 5 PO-FW26-001 SKUs (7,800 units) and all 3 PO-FW26-002 SKUs "
            "(3,000 units) at exact expected quantities. "
            "2 POs in the summary block. Zero over/under. "
            "All CI SKUs match their respective PO line items."
        ),
        "pos": ["PO-FW26-001", "PO-FW26-002"],
        "skus": [
            ("TEN7101-FGN-XS",   800, "matched — PO-FW26-001"),
            ("TEN7101-FGN-S",   1400, "matched — PO-FW26-001"),
            ("TEN7101-FGN-M",   2000, "matched — PO-FW26-001"),
            ("TEN7101-FGN-L",   2000, "matched — PO-FW26-001"),
            ("TEN7101-FGN-XL",  1600, "matched — PO-FW26-001"),
            ("TEN7102-NVY-XS",   500, "matched — PO-FW26-002"),
            ("TEN7102-NVY-S",   1000, "matched — PO-FW26-002"),
            ("TEN7102-NVY-M",   1500, "matched — PO-FW26-002"),
        ],
        "parse": {
            "total_shipped_qty": 10800,
            "matched_skus":      8,
            "unmatched_skus":    0,
            "over_shipped_skus": 0,
        },
    },
))


# ─────────────────────────────────────────────────────────────────────────────
# S3 Multi-PO Mixed: partial + unmatched across both POs
# ─────────────────────────────────────────────────────────────────────────────
s3mp_items = [
    li("TEN7101-FGN-XS",    "FW26 Fleece Jacket — Forest Green XS",  800,  16.50, 0.20, 0.0025),  # PO-001 matched
    li("TEN7101-FGN-S",     "FW26 Fleece Jacket — Forest Green S",  1400,  16.50, 0.20, 0.0025),  # PO-001 matched
    li("TEN7101-FGN-M",     "FW26 Fleece Jacket — Forest Green M",  2000,  16.50, 0.20, 0.0025),  # PO-001 matched
    li("TEN7102-NVY-XS",    "FW26 Hoodie — Navy XS",                 500,  18.00, 0.22, 0.0028),  # PO-002 matched
    li("TEN7102-NVY-GHOST", "FW26 Hoodie — Navy GHOST",              500,  18.00, 0.22, 0.0028),  # unmatched on PO-002
    li("TEN7101-EXTRA",     "FW26 Extra Style",                      300,  16.50, 0.20, 0.0025),  # unmatched on any PO
]

s3mp_wb = build_ci_workbook(
    invoice_number  = "CI-S3MP-MIX-001",
    invoice_date    = INVOICE_DATE,
    po_summary_rows = [
        ("PO-FW26-001", 4200, 210, 840.0, 10.50),
        ("PO-FW26-002", 1000,  50, 220.0,  2.80),
    ],
    line_items = s3mp_items,
)
scenarios.append((
    "ci_s3_multi_po_mixed.xlsx",
    s3mp_wb,
    {
        "title": "S3 Multi-PO Mixed — Partial + Unmatched (PO-FW26-001 + PO-FW26-002)",
        "desc": (
            "3 matched SKUs from PO-FW26-001 (XS+S+M = 4,200 units). "
            "1 matched SKU from PO-FW26-002 (NVY-XS = 500 units). "
            "1 unmatched on PO-002 (TEN7102-NVY-GHOST = 500). "
            "1 fully unmatched across all POs (TEN7101-EXTRA = 300). "
            "L and XL from PO-001 have 0 shipped. "
            "NVY-S and NVY-M from PO-002 have 0 shipped."
        ),
        "pos": ["PO-FW26-001", "PO-FW26-002"],
        "skus": [
            ("TEN7101-FGN-XS",     800, "matched — PO-FW26-001"),
            ("TEN7101-FGN-S",     1400, "matched — PO-FW26-001"),
            ("TEN7101-FGN-M",     2000, "matched — PO-FW26-001"),
            ("TEN7102-NVY-XS",    500,  "matched — PO-FW26-002"),
            ("TEN7102-NVY-GHOST", 500,  "UNMATCHED — not on PO-FW26-002"),
            ("TEN7101-EXTRA",     300,  "UNMATCHED — not on any PO"),
        ],
        "parse": {
            "total_shipped_qty": 5500,
            "matched_skus":      4,
            "unmatched_skus":    2,
            "over_shipped_skus": 0,
        },
    },
))


# ── 6. Save all workbooks ─────────────────────────────────────────────────────

print(f"\nWriting CI fixture files to: {DATA_DIR}\n")
saved = []
for filename, wb, _ in scenarios:
    out_path = DATA_DIR / filename
    wb.save(str(out_path))
    size_kb = out_path.stat().st_size / 1024
    print(f"  [OK] {filename:48s}  ({size_kb:.1f} KB)")
    saved.append(filename)

print(f"\n{len(saved)} / 7 files written.\n")


# ── 7. Write CI_FIXTURES.md ───────────────────────────────────────────────────

def sku_table_md(skus):
    rows = ["| SKU Code | Qty | Status |", "|---|---:|---|"]
    for sku, qty, status in skus:
        rows.append(f"| `{sku}` | {qty:,} | {status} |")
    return "\n".join(rows)


md_lines = [
    "# CI Fixture Files — Test Scenarios",
    "",
    "Generated by `backend/scripts/create_ci_fixtures.py`.",
    "All files live in `backend/data/`. Invoice date for all fixtures: **2026-04-15**.",
    "",
    "---",
    "",
]

for filename, _, meta in scenarios:
    ep = meta["parse"]
    md_lines += [
        f"## `{filename}`",
        "",
        f"### {meta['title']}",
        "",
        meta["desc"],
        "",
        f"**POs in summary block:** {', '.join(f'`{p}`' for p in meta['pos'])}",
        "",
        "### SKU Line Items",
        "",
        sku_table_md(meta["skus"]),
        "",
        "### Expected Parse Outcomes",
        "",
        "| Metric | Value |",
        "|---|---:|",
        f"| Total CI shipped qty | {ep['total_shipped_qty']:,} |",
        f"| Matched SKUs | {ep['matched_skus']} |",
        f"| Unmatched SKUs (not on any PO) | {ep['unmatched_skus']} |",
        f"| Over-shipped SKUs | {ep['over_shipped_skus']} |",
        "",
        "---",
        "",
    ]

md_path = DATA_DIR / "CI_FIXTURES.md"
md_path.write_text("\n".join(md_lines), encoding="utf-8")
print(f"  [OK] CI_FIXTURES.md written.\n")
print("Done.\n")
